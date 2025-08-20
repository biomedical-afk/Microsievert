import streamlit as st
import pandas as pd
import requests
from io import BytesIO
from datetime import datetime
from dateutil.parser import parse as dtparse
from typing import List, Dict, Any, Optional, Set

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.drawing.image import Image as XLImage

# Logo de ejemplo si no subes uno (opcional)
try:
    from PIL import Image as PILImage, ImageDraw, ImageFont
except Exception:
    PILImage = None

# ================== CREDENCIALES NINOX ==================
API_TOKEN   = "0b3a1130-785a-11f0-ace0-3fb1fcb242e2"  # <-- tu API key
TEAM_ID     = "ihp8o8AaLzfodwc4J"
DATABASE_ID = "ksqzvuts5aq0"
BASE_URL    = "https://api.ninox.com/v1"
TABLE_ID    = "C"   # Tabla REPORTE
# ========================================================

st.set_page_config(page_title="Reporte de Dosimetría — Ninox", layout="wide")
st.title("Reporte de Dosimetría — Actual, Anual y de por Vida (por persona)")

# ---------------------- Utilidades ----------------------
def headers() -> Dict[str, str]:
    return {"Authorization": f"Bearer {API_TOKEN}", "Content-Type": "application/json"}

def as_value(v: Any):
    """Para mostrar: conserva 'PM' y convierte números (coma→punto)."""
    if v is None: return ""
    s = str(v).strip().replace(",", ".")
    if s.upper() == "PM": return "PM"
    try: return float(s)
    except Exception: return s

def as_num(v: Any) -> float:
    """Para sumas: PM/vacío -> 0.0."""
    if v is None: return 0.0
    s = str(v).strip().replace(",", ".")
    if s == "" or s.upper() == "PM": return 0.0
    try: return float(s)
    except Exception: return 0.0

def round2(x: float) -> float:
    return float(f"{x:.2f}")

def fetch_all_records(table_id: str, page_size: int = 1000):
    url = f"{BASE_URL}/teams/{TEAM_ID}/databases/{DATABASE_ID}/tables/{table_id}/records"
    skip, out = 0, []
    while True:
        r = requests.get(url, headers=headers(), params={"limit": page_size, "skip": skip}, timeout=60)
        r.raise_for_status()
        chunk = r.json()
        if not chunk: break
        out.extend(chunk)
        if len(chunk) < page_size: break
        skip += page_size
    return out

def normalize_df(records):
    rows = []
    for r in records:
        f = r.get("fields", {}) or {}
        rows.append({
            "_id": r.get("id"),
            "PERIODO DE LECTURA": f.get("PERIODO DE LECTURA"),
            "COMPAÑÍA": f.get("COMPAÑÍA"),
            "CÓDIGO DE DOSÍMETRO": str(f.get("CÓDIGO DE DOSÍMETRO") or "").strip(),
            "NOMBRE": f.get("NOMBRE"),
            "CÉDULA": f.get("CÉDULA"),
            "FECHA DE LECTURA": f.get("FECHA DE LECTURA"),
            "TIPO DE DOSÍMETRO": f.get("TIPO DE DOSÍMETRO"),
            # RAW (mostrar, conserva PM) y NUM (sumas)
            "Hp10_RAW":  as_value(f.get("Hp (10)")),
            "Hp007_RAW": as_value(f.get("Hp (0.07)")),
            "Hp3_RAW":  as_value(f.get("Hp (3)")),
            "Hp10_NUM":  as_num(f.get("Hp (10)")),
            "Hp007_NUM": as_num(f.get("Hp (0.07)")),
            "Hp3_NUM":  as_num(f.get("Hp (3)")),
        })
    df = pd.DataFrame(rows)
    df["FECHA_DE_LECTURA_DT"] = pd.to_datetime(
        df["FECHA DE LECTURA"].apply(
            lambda x: dtparse(str(x), dayfirst=True) if pd.notna(x) and str(x).strip() != "" else pd.NaT
        ), errors="coerce"
    )
    # Claves de persona
    df["NOMBRE_NORM"] = df["NOMBRE"].fillna("").astype(str).str.strip()
    df["CÉDULA_NORM"] = df["CÉDULA"].fillna("").astype(str).str.strip()
    return df

def read_codes_from_files(files) -> Set[str]:
    """Lee CSV/Excel y extrae códigos (columna candidata o patrón WB\\d+)."""
    codes: Set[str] = set()
    from io import BytesIO
    for f in files:
        raw = f.read(); f.seek(0)
        name = f.name.lower()
        try:
            if name.endswith((".xlsx", ".xls")):
                df = pd.read_excel(BytesIO(raw))
            else:
                df = None
                for enc in ("utf-8-sig","latin-1"):
                    try:
                        df = pd.read_csv(BytesIO(raw), sep=None, engine="python", encoding=enc); break
                    except Exception: continue
                if df is None: df = pd.read_csv(BytesIO(raw))
        except Exception:
            continue
        if df is None or df.empty: continue
        cand = None
        for c in df.columns:
            cl = str(c).lower()
            if any(k in cl for k in ["dosim","código","codigo","wb","dosímetro","dosimetro"]):
                cand = c; break
        if cand is None:
            for c in df.columns:
                if df[c].astype(str).str.contains(r"^WB\d{5,}$", case=False, na=False).any():
                    cand = c; break
        if cand is None: cand = df.columns[0]
        codes |= set(df[cand].astype(str).str.strip())
    return {c for c in codes if c and c.lower() != "nan"}

# --- Helpers robustos para PM/listas ---
def pm_or_sum(raws, numeric_sum) -> Any:
    """
    Acepta raws como lista/tupla/set/Series/un valor único/NaN.
    Si TODAS las lecturas válidas son 'PM' -> 'PM'.
    Si no, devuelve la suma numérica redondeada a 2 decimales.
    """
    import pandas as _pd

    # Normalizar raws a lista
    if isinstance(raws, (list, tuple, set)):
        arr = list(raws)
    elif isinstance(raws, _pd.Series):
        arr = raws.tolist()
    elif raws is None or (isinstance(raws, float) and _pd.isna(raws)) or raws == "":
        arr = []
    else:
        arr = [raws]

    vals = [str(x).upper() for x in arr if str(x).strip() != ""]
    if vals and all(v == "PM" for v in vals):
        return "PM"

    # Proteger numeric_sum
    try:
        total = float(numeric_sum)
        if _pd.isna(total):
            total = 0.0
    except Exception:
        total = 0.0
    return round2(total)

def merge_raw_lists(*vals):
    """Combina varias 'listas' de raws tolerando NaN/None/valores sueltos."""
    import pandas as _pd
    out: List[Any] = []
    for v in vals:
        if isinstance(v, (list, tuple, set)):
            out.extend(list(v))
        elif isinstance(v, _pd.Series):
            out.extend(v.tolist())
        elif v is None or (isinstance(v, float) and _pd.isna(v)) or v == "":
            continue
        else:
            out.append(v)
    return out

# ------------------- Sidebar -------------------
with st.sidebar:
    st.header("Filtros")
    files = st.file_uploader("Archivos de dosis (para filtrar por CÓDIGO DE DOSÍMETRO)", type=["csv","xlsx","xls"], accept_multiple_files=True)

    st.markdown("---")
    st.subheader("Encabezado del Excel")
    header_line1 = st.text_input("Línea 1", "MICROSIEVERT, S.A.")
    header_line2 = st.text_input("Línea 2", "PH Conardo")
    header_line3 = st.text_input("Línea 3", "Calle 41 Este, Panamá")
    header_line4 = st.text_input("Línea 4", "PANAMÁ")
    logo_file = st.file_uploader("Logo (PNG/JPG) opcional", type=["png","jpg","jpeg"])

# ------------------- Carga Ninox -------------------
with st.spinner("Cargando datos desde Ninox…"):
    base = normalize_df(fetch_all_records(TABLE_ID))

if base.empty:
    st.warning("No hay registros en la tabla REPORTE.")
    st.stop()

# ------------------- Filtros adicionales -------------------
with st.sidebar:
    st.markdown("---")
    per_order = (base.groupby("PERIODO DE LECTURA")["FECHA_DE_LECTURA_DT"].max()
                 .sort_values(ascending=False).index.astype(str).tolist())
    per_valid = [p for p in per_order if p.strip().upper() != "CONTROL"]
    periodo_actual = st.selectbox("Periodo actual", per_valid, index=0 if per_valid else None)

    comp_opts = ["(todas)"] + sorted(base["COMPAÑÍA"].dropna().astype(str).unique().tolist())
    compania = st.selectbox("Compañía", comp_opts, index=0)
    tipo_opts = ["(todos)"] + sorted(base["TIPO DE DOSÍMETRO"].dropna().astype(str).unique().tolist())
    tipo = st.selectbox("Tipo de dosímetro", tipo_opts, index=0)

# Filtro por archivos / compañía / tipo
codes_filter: Optional[Set[str]] = None
if files:
    codes_filter = read_codes_from_files(files)
    if codes_filter:
        st.success(f"Códigos detectados: {len(codes_filter)}")

df = base.copy()
if codes_filter:
    df = df[df["CÓDIGO DE DOSÍMETRO"].isin(codes_filter)]
if compania != "(todas)":
    df = df[df["COMPAÑÍA"].astype(str) == compania]
if tipo != "(todos)":
    df = df[df["TIPO DE DOSÍMETRO"].astype(str) == tipo]

if df.empty:
    st.warning("No hay registros que cumplan el filtro.")
    st.stop()

# === EXCLUIR CONTROL DE LOS CÁLCULOS (pero luego lo mostraremos primero si existe) ===
df["__is_control_row"] = df["NOMBRE"].fillna("").astype(str).str.strip().str.upper().eq("CONTROL")
df = df[~df["__is_control_row"]].copy()

# ------------------- Cálculos POR PERSONA (NOMBRE + CÉDULA) -------------------
keys = ["NOMBRE_NORM", "CÉDULA_NORM"]

# 0) Determinar el año objetivo a partir del 'periodo_actual'
df_periodo = df[df["PERIODO DE LECTURA"].astype(str) == str(periodo_actual)].copy()
if df_periodo.empty:
    st.warning("No hay filas en el periodo actual seleccionado.")
    st.stop()
year_target = df_periodo["FECHA_DE_LECTURA_DT"].dt.year.max()

# 1) ACTUAL: sumar SOLO el periodo_actual por persona
df_curr = df[df["PERIODO DE LECTURA"].astype(str) == str(periodo_actual)].copy()
df_curr = df_curr.sort_values("FECHA_DE_LECTURA_DT")  # para que 'last' sea el más reciente
gb_curr_sum = df_curr.groupby(keys, as_index=False).agg({
    "PERIODO DE LECTURA": "last",
    "COMPAÑÍA": "last",
    "CÓDIGO DE DOSÍMETRO": "last",
    "NOMBRE": "last",
    "CÉDULA": "last",
    "FECHA_DE_LECTURA_DT": "max",
    "TIPO DE DOSÍMETRO": "last",
    "Hp10_NUM": "sum",
    "Hp007_NUM": "sum",
    "Hp3_NUM": "sum",
})
gb_curr_raw = df_curr.groupby(keys).agg({
    "Hp10_RAW": list,
    "Hp007_RAW": list,
    "Hp3_RAW": list
}).rename(columns={
    "Hp10_RAW": "Hp10_ACTUAL_RAW_LIST",
    "Hp007_RAW": "Hp007_ACTUAL_RAW_LIST",
    "Hp3_RAW": "Hp3_ACTUAL_RAW_LIST"
}).reset_index()

out = gb_curr_sum.merge(gb_curr_raw, on=keys, how="left").rename(columns={
    "Hp10_NUM": "Hp10_ACTUAL_NUM_SUM",
    "Hp007_NUM": "Hp007_ACTUAL_NUM_SUM",
    "Hp3_NUM": "Hp3_ACTUAL_NUM_SUM",
})

# 2) ANUAL (automático): todas las filas del MISMO AÑO del periodo actual
df_year = df[df["FECHA_DE_LECTURA_DT"].dt.year == year_target].copy()

gb_year_sum = df_year.groupby(keys, as_index=False).agg({
    "Hp10_NUM": "sum",
    "Hp007_NUM": "sum",
    "Hp3_NUM": "sum"
}).rename(columns={
    "Hp10_NUM": "Hp10_YEAR_NUM_SUM",
    "Hp007_NUM": "Hp007_YEAR_NUM_SUM",
    "Hp3_NUM": "Hp3_YEAR_NUM_SUM",
})

gb_year_raw = df_year.groupby(keys).agg({
    "Hp10_RAW": list,
    "Hp007_RAW": list,
    "Hp3_RAW": list
}).rename(columns={
    "Hp10_RAW": "Hp10_YEAR_RAW_LIST",
    "Hp007_RAW": "Hp007_YEAR_RAW_LIST",
    "Hp3_RAW": "Hp3_YEAR_RAW_LIST"
}).reset_index()

out = out.merge(gb_year_sum, on=keys, how="left").merge(gb_year_raw, on=keys, how="left")

# 3) VIDA (todo el historial de la persona)
gb_life_sum = df.groupby(keys, as_index=False).agg({
    "Hp10_NUM": "sum",
    "Hp007_NUM": "sum",
    "Hp3_NUM": "sum"
}).rename(columns={
    "Hp10_NUM": "Hp10_LIFE_NUM_SUM",
    "Hp007_NUM": "Hp007_LIFE_NUM_SUM",
    "Hp3_NUM": "Hp3_LIFE_NUM_SUM",
})

gb_life_raw = df.groupby(keys).agg({
    "Hp10_RAW": list,
    "Hp007_RAW": list,
    "Hp3_RAW": list
}).rename(columns={
    "Hp10_RAW": "Hp10_LIFE_RAW_LIST",
    "Hp007_RAW": "Hp007_LIFE_RAW_LIST",
    "Hp3_RAW": "Hp3_LIFE_RAW_LIST"
}).reset_index()

out = out.merge(gb_life_sum, on=keys, how="left").merge(gb_life_raw, on=keys, how="left")

# Rellenos numéricos
for c in [
    "Hp10_ACTUAL_NUM_SUM","Hp007_ACTUAL_NUM_SUM","Hp3_ACTUAL_NUM_SUM",
    "Hp10_YEAR_NUM_SUM","Hp007_YEAR_NUM_SUM","Hp3_YEAR_NUM_SUM",
    "Hp10_LIFE_NUM_SUM","Hp007_LIFE_NUM_SUM","Hp3_LIFE_NUM_SUM",
]:
    if c not in out.columns: out[c] = 0.0
    out[c] = out[c].fillna(0.0)

# ---- Columnas de salida (PM donde corresponde) ----
def fmt_fecha(dtval):
    if pd.isna(dtval): return ""
    try: return pd.to_datetime(dtval).strftime("%d/%m/%Y %H:%M")
    except Exception: return str(dtval)

# ACTUAL
out["Hp (10) ACTUAL"]   = out.apply(lambda r: pm_or_sum(r.get("Hp10_ACTUAL_RAW_LIST", []), r["Hp10_ACTUAL_NUM_SUM"]), axis=1)
out["Hp (0.07) ACTUAL"] = out.apply(lambda r: pm_or_sum(r.get("Hp007_ACTUAL_RAW_LIST", []), r["Hp007_ACTUAL_NUM_SUM"]), axis=1)
out["Hp (3) ACTUAL"]    = out.apply(lambda r: pm_or_sum(r.get("Hp3_ACTUAL_RAW_LIST",  []), r["Hp3_ACTUAL_NUM_SUM"]),  axis=1)

# ANUAL (año del periodo actual)
out["Hp (10) ANUAL"] = out.apply(
    lambda r: pm_or_sum(r.get("Hp10_YEAR_RAW_LIST", []), float(r["Hp10_YEAR_NUM_SUM"])), axis=1
)
out["Hp (0.07) ANUAL"] = out.apply(
    lambda r: pm_or_sum(r.get("Hp007_YEAR_RAW_LIST", []), float(r["Hp007_YEAR_NUM_SUM"])), axis=1
)
out["Hp (3) ANUAL"] = out.apply(
    lambda r: pm_or_sum(r.get("Hp3_YEAR_RAW_LIST", []), float(r["Hp3_YEAR_NUM_SUM"])), axis=1
)

# VIDA (todo el historial)
out["Hp (10) VIDA"]   = out.apply(lambda r: pm_or_sum(r.get("Hp10_LIFE_RAW_LIST", []), r["Hp10_LIFE_NUM_SUM"]), axis=1)
out["Hp (0.07) VIDA"] = out.apply(lambda r: pm_or_sum(r.get("Hp007_LIFE_RAW_LIST", []), r["Hp007_LIFE_NUM_SUM"]), axis=1)
out["Hp (3) VIDA"]    = out.apply(lambda r: pm_or_sum(r.get("Hp3_LIFE_RAW_LIST",  []), r["Hp3_LIFE_NUM_SUM"]),  axis=1)

# columnas de fecha y periodo
out["FECHA Y HORA DE LECTURA"] = out["FECHA_DE_LECTURA_DT"].apply(fmt_fecha)
out["PERIODO DE LECTURA"] = periodo_actual

# Orden (CONTROL al inicio si existe)
out["__is_control"] = out["NOMBRE"].fillna("").astype(str).str.strip().str.upper().eq("CONTROL")
out = out.sort_values(["__is_control","NOMBRE","CÉDULA"], ascending=[False, True, True])

# ---- Columnas finales ----
FINAL_COLS = [
    "PERIODO DE LECTURA","COMPAÑÍA","CÓDIGO DE DOSÍMETRO","NOMBRE","CÉDULA",
    "FECHA Y HORA DE LECTURA","TIPO DE DOSÍMETRO",
    "Hp (10) ACTUAL","Hp (0.07) ACTUAL","Hp (3) ACTUAL",
    "Hp (10) ANUAL","Hp (0.07) ANUAL","Hp (3) ANUAL",
    "Hp (10) VIDA","Hp (0.07) VIDA","Hp (3) VIDA",
]
for c in FINAL_COLS:
    if c not in out.columns: out[c] = ""
out = out[FINAL_COLS]

# ------------------- Vista previa -------------------
st.subheader("Reporte final (vista previa)")
st.dataframe(out, use_container_width=True, hide_index=True)

# ------------------- Descargas -------------------
# CSV
csv_bytes = out.to_csv(index=False).encode("utf-8-sig")
st.download_button(
    "⬇️ Descargar CSV (UTF-8 con BOM)",
    data=csv_bytes,
    file_name=f"reporte_dosimetria_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
    mime="text/csv"
)

# Excel simple
def to_excel_simple(df: pd.DataFrame, sheet_name="Reporte"):
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name=sheet_name)
    bio.seek(0); return bio.getvalue()

xlsx_simple = to_excel_simple(out)
st.download_button(
    "⬇️ Descargar Excel (tabla simple)",
    data=xlsx_simple,
    file_name=f"reporte_dosimetria_tabla_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

# ---------- Helpers de Excel (logo y medidas) ----------
def col_pixels(ws, col_letter: str) -> int:
    w = ws.column_dimensions[col_letter].width
    if w is None: w = 8.43
    return int(w * 7 + 5)

def row_pixels(ws, row_idx: int) -> int:
    h = ws.row_dimensions[row_idx].height
    if h is None: h = 15  # pt por defecto
    return int(h * 96 / 72)  # 1pt = 1/72", 96dpi aprox.

def fit_logo(ws, logo_bytes: bytes, top_left: str = "C1", bottom_right: str = "F4", padding: int = 6):
    """Escala y coloca el logo dentro del rectángulo top_left..bottom_right conservando proporción."""
    if not logo_bytes: return
    img = XLImage(BytesIO(logo_bytes))

    tl_col = column_index_from_string(''.join([c for c in top_left if c.isalpha()]))
    tl_row = int(''.join([c for c in top_left if c.isdigit()]))
    br_col = column_index_from_string(''.join([c for c in bottom_right if c.isalpha()]))
    br_row = int(''.join([c for c in bottom_right if c.isdigit()]))

    box_w = sum(col_pixels(ws, get_column_letter(c)) for c in range(tl_col, br_col + 1))
    box_h = sum(row_pixels(ws, r) for r in range(tl_row, br_row + 1))

    max_w = max(10, box_w - 2*padding)
    max_h = max(10, box_h - 2*padding)

    scale = min(max_w / img.width, max_h / img.height, 1.0)
    img.width = int(img.width * scale)
    img.height = int(img.height * scale)
    img.anchor = top_left
    ws.add_image(img)

def sample_logo_bytes(text="µSv  MICROSIEVERT, S.A."):
    """Crea un logo de ejemplo si no subes uno (usa Pillow)."""
    if PILImage is None: return None
    img = PILImage.new("RGBA", (420, 110), (255, 255, 255, 0))
    d = ImageDraw.Draw(img)
    try:
        font = ImageFont.truetype("arial.ttf", 36)
    except Exception:
        font = ImageFont.load_default()
    d.text((12, 30), text, fill=(0, 70, 140, 255), font=font)
    bio = BytesIO(); img.save(bio, format="PNG"); return bio.getvalue()

# ------------------- Excel formato plantilla -------------------
def build_formatted_excel(df_final: pd.DataFrame,
                          header_lines: List[str],
                          logo_bytes: Optional[bytes]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    gray = PatternFill("solid", fgColor="DDDDDD")
    group_fill = PatternFill("solid", fgColor="EEEEEE")

    # Anchos/altos cabecera (para el cuadro del logo)
    widths = {
        "A": 24, "B": 28, "C": 16, "D": 16, "E": 16, "F": 16,
        "G": 10, "H": 12, "I": 12, "J": 12, "K": 12, "L": 12, "M": 12,
        "N": 12, "O": 12, "P": 12
    }
    for k, v in widths.items():
        ws.column_dimensions[k].width = v
    for r in range(1, 4 + 1):
        ws.row_dimensions[r].height = 20  # altura homogénea para el área del logo

    # Encabezado texto (A1:B4)
    for i, line in enumerate(header_lines[:4], start=1):
        ws.merge_cells(f"A{i}:B{i}")
        c = ws[f"A{i}"]; c.value = line; c.fill = gray
        c.font = Font(bold=True); c.alignment = Alignment(horizontal="left", vertical="center")
        for col in ("A","B"): ws.cell(row=i, column=ord(col)-64).border = border

    # Fecha de emisión (I1:P1)
    ws.merge_cells("I1:J1"); ws["I1"] = "Fecha de emisión"
    ws["I1"].font = Font(bold=True, size=10); ws["I1"].alignment = center; ws["I1"].fill = gray
    ws.merge_cells("K1:P1"); ws["K1"] = datetime.now().strftime("%d-%b-%y").lower()
    ws["K1"].font = Font(bold=True, size=10); ws["K1"].alignment = center
    for col_idx in range(ord("I")-64, ord("P")-64+1):
        ws.cell(row=1, column=col_idx).border = border

    # Logo en C1:F4 (escala automática)
    if logo_bytes is None:
        logo_bytes = sample_logo_bytes()
    if logo_bytes:
        fit_logo(ws, logo_bytes, top_left="C1", bottom_right="F4", padding=6)

    # Título
    ws.merge_cells("A6:P6")
    ws["A6"] = "REPORTE DE DOSIMETRÍA"
    ws["A6"].font = Font(bold=True, size=14)
    ws["A6"].alignment = center

    # Bloques (cerrados con borde exactamente sobre sus columnas)
    ws.merge_cells("H7:J7"); ws["H7"] = "DOSIS ACTUAL (mSv)"
    ws.merge_cells("K7:M7"); ws["K7"] = "DOSIS ANUAL (mSv)"
    ws.merge_cells("N7:P7"); ws["N7"] = "DOSIS DE POR VIDA (mSv)"
    for rng in (("H7","J7"), ("K7","M7"), ("N7","P7")):
        start_col = column_index_from_string(rng[0][0]); end_col = column_index_from_string(rng[1][0])
        row = 7
        ws[rng[0]].font = bold; ws[rng[0]].alignment = center; ws[rng[0]].fill = group_fill
        for col in range(start_col, end_col + 1):
            c = ws.cell(row=row, column=col)
            c.border = border; c.fill = group_fill

    # Encabezados de tabla
    headers = [
        "PERIODO DE LECTURA","COMPAÑÍA","CÓDIGO DE DOSÍMETRO","NOMBRE","CÉDULA",
        "FECHA Y HORA DE LECTURA","TIPO DE DOSÍMETRO",
        "Hp (10) ACTUAL","Hp (0.07) ACTUAL","Hp (3) ACTUAL",
        "Hp (10) ANUAL","Hp (0.07) ANUAL","Hp (3) ANUAL",
        "Hp (10) VIDA","Hp (0.07) VIDA","Hp (3) VIDA",
    ]
    header_row = 8
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = bold; cell.alignment = center; cell.border = border
        cell.fill = gray

    # Datos
    start_row = header_row + 1
    for _, r in df_final[headers].iterrows():
        ws.append(list(r.values))
    last_row = ws.max_row

    # Bordes, alineación y altura uniforme
    for row in ws.iter_rows(min_row=header_row, max_row=last_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = border
            if cell.row >= start_row:
                cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    for rr in range(start_row, last_row + 1):
        ws.row_dimensions[rr].height = 20

    ws.freeze_panes = f"A{start_row}"

    # Auto ancho (con máximos)
    for col_cells in ws.iter_cols(min_col=1, max_col=16, min_row=header_row, max_row=last_row):
        col_letter = get_column_letter(col_cells[0].column)
        max_len = max(len("" if c.value is None else str(c.value)) for c in col_cells)
        ws.column_dimensions[col_letter].width = max(ws.column_dimensions[col_letter].width, min(max_len + 2, 42))

    # ------------------ Sección informativa ------------------
    row = last_row + 2
    ws.merge_cells(f"A{row}:P{row}")
    ws[f"A{row}"] = "INFORMACIÓN DEL REPORTE DE DOSIMETRÍA"
    ws[f"A{row}"].font = Font(bold=True); ws[f"A{row}"].alignment = Alignment(horizontal="center")
    row += 1

    bullets = [
        "‒ Periodo de lectura: periodo de uso del dosímetro personal.",
        "‒ Fecha de lectura: fecha en que se realizó la lectura.",
        "‒ Tipo de dosímetro:",
    ]
    for text in bullets:
        ws.merge_cells(f"A{row}:D{row}")
        c = ws[f"A{row}"]; c.value = text
        c.font = Font(size=10, bold=True); c.alignment = Alignment(horizontal="left")
        row += 2

    tipos = [("CE","Cuerpo Entero"), ("A","Anillo"), ("B","Brazalete"), ("CR","Cristalino")]
    for clave, desc in tipos:
        ws.merge_cells(f"C{row}:D{row}")
        c = ws[f"C{row}"]; c.value = f"{clave} = {desc}"
        c.font = Font(size=10, bold=True); c.alignment = Alignment(horizontal="left")
        for col in ("C","D"): ws.cell(row=row, column=ord(col)-64).border = border
        row += 1
    row += 1

    ws.merge_cells(f"F{row}:I{row}")
    ws[f"F{row}"] = "LÍMITES ANUALES DE EXPOSICIÓN A RADIACIONES"
    ws[f"F{row}"].font = Font(bold=True, size=10); ws[f"F{row}"].alignment = Alignment(horizontal="center")
    row += 1

    limites = [
        ("Cuerpo Entero", "20 mSv/año"),
        ("Cristalino", "150 mSv/año"),
        ("Extremidades y piel", "500 mSv/año"),
        ("Fetal", "1 mSv/periodo de gestación"),
        ("Público", "1 mSv/año"),
    ]
    for cat, val in limites:
        ws.merge_cells(f"F{row}:G{row}"); ws[f"F{row}"].value = cat
        ws[f"F{row}"].font = Font(size=10); ws[f"F{row}"].alignment = Alignment(horizontal="left")
        ws.merge_cells(f"H{row}:I{row}"); ws[f"H{row}"].value = val
        ws[f"H{row}"].font = Font(size=10); ws[f"H{row}"].alignment = Alignment(horizontal="right")
        for col in ("F","G","H","I"):
            ws.cell(row=row, column=ord(col)-64).border = border
        row += 1
    row += 2

    ws.merge_cells(f"A{row}:P{row}")
    ws[f"A{row}"] = "‒ DATOS DEL PARTICIPANTE:"
    ws[f"A{row}"].font = Font(bold=True, size=10); ws[f"A{row}"].alignment = Alignment(horizontal="left")
    row += 1

    datos = [
        "‒ Código de usuario: Número único asignado al usuario por Microsievert, S.A.",
        "‒ Nombre: Persona a la cual se le asigna el dosímetro personal.",
        "‒ Cédula: Número del documento de identidad personal del usuario.",
    ]
    for txt in datos:
        ws.merge_cells(f"A{row}:P{row}")
        ws[f"A{row}"].value = txt
        ws[f"A{row}"].font = Font(size=10)
        ws[f"A{row}"].alignment = Alignment(horizontal="left")
        row += 1
    row += 2

    ws.merge_cells(f"A{row}:P{row}")
    ws[f"A{row}"] = "‒ DOSIS EN MILISIEVERT:"
    ws[f"A{row}"].font = Font(bold=True, size=10); ws[f"A{row}"].alignment = Alignment(horizontal="left")
    row += 1

    shade = PatternFill("solid", fgColor="DDDDDD")
    ws.merge_cells(f"B{row}:C{row}"); ws[f"B{row}"] = "Nombre"
    ws[f"B{row}"].font = Font(bold=True, size=10)
    ws[f"B{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); ws[f"B{row}"].fill = shade
    ws.merge_cells(f"D{row}:I{row}"); ws[f"D{row}"] = "Definición"
    ws[f"D{row}"].font = Font(bold=True, size=10)
    ws[f"D{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); ws[f"D{row}"].fill = shade
    ws.merge_cells(f"J{row}:J{row}"); ws[f"J{row}"] = "Unidad"
    ws[f"J{row}"].font = Font(bold=True, size=10)
    ws[f"J{row}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); ws[f"J{row}"].fill = shade
    for col in ("B","C","D","E","F","G","H","I","J"):
        ws.cell(row=row, column=ord(col)-64).border = border
    ws.row_dimensions[row].height = 30
    row += 1

    definitions = [
        ("Dosis efectiva Hp(10)",  "Es la dosis equivalente en tejido blando, J·kg⁻¹ o Sv a una profundidad de 10 mm, bajo determinado punto del cuerpo.", "mSv"),
        ("Dosis superficial Hp(0,07)", "Es la dosis equivalente en tejido blando, J·kg⁻¹ o Sv a una profundidad de 0,07 mm, bajo determinado punto del cuerpo.", "mSv"),
        ("Dosis cristalino Hp(3)", "Es la dosis equivalente en tejido blando, J·kg⁻¹ o Sv a una profundidad de 3 mm, bajo determinado punto del cuerpo.", "mSv"),
    ]
    for nom, desc, uni in definitions:
        ws.merge_cells(f"B{row}:C{row}"); ws[f"B{row}"] = nom
        ws[f"B{row}"].font = Font(size=10, bold=True); ws[f"B{row}"].alignment = Alignment(horizontal="left", wrap_text=True)
        ws.merge_cells(f"D{row}:I{row}"); ws[f"D{row}"] = desc
        ws[f"D{row}"].font = Font(size=10); ws[f"D{row}"].alignment = Alignment(horizontal="left", wrap_text=True)
        ws.merge_cells(f"J{row}:J{row}"); ws[f"J{row}"] = uni
        ws[f"J{row}"].font = Font(size=10); ws[f"J{row}"].alignment = Alignment(horizontal="center", wrap_text=True)
        for col in ("B","C","D","E","F","G","H","I","J"):
            ws.cell(row=row, column=ord(col)-64).border = border
        ws.row_dimensions[row].height = 30
        row += 1

    row += 1
    ws.merge_cells(f"A{row}:P{row}")
    ws[f"A{row}"] = "LECTURAS DE ANILLO: las lecturas del dosímetro de anillo son registradas como una dosis equivalente superficial Hp(0,07)."
    ws[f"A{row}"].font = Font(size=10, bold=True); ws[f"A{row}"].alignment = Alignment(horizontal="left", wrap_text=True)
    row += 1

    ws.merge_cells(f"A{row}:P{row}")
    ws[f"A{row}"] = "Los resultados de las dosis individuales de radiación son reportados para diferentes periodos de tiempo:"
    ws[f"A{row}"].font = Font(size=10); ws[f"A{row}"].alignment = Alignment(horizontal="left", wrap_text=True)
    row += 1

    blocks = [
        ("DOSIS ACTUAL",      "Es el correspondiente de dosis acumulada durante el período de lectura definido."),
        ("DOSIS ANUAL",       "Es el correspondiente de dosis acumulada desde el inicio del año hasta la fecha."),
        ("DOSIS DE POR VIDA", "Es el correspondiente de dosis acumulada desde el inicio del servicio dosimétrico hasta la fecha."),
    ]
    for clave, texto in blocks:
        ws.merge_cells(f"B{row}:C{row}"); ws[f"B{row}"] = clave
        ws[f"B{row}"].font = Font(bold=True, size=10); ws[f"B{row}"].alignment = Alignment(horizontal="center")
        ws.merge_cells(f"D{row}:P{row}"); ws[f"D{row}"] = texto
        ws[f"D{row}"].font = Font(size=10); ws[f"D{row}"].alignment = Alignment(horizontal="left", wrap_text=True)
        for col_idx in range(ord("B")-64, ord("P")-64+1):
            ws.cell(row=row, column=col_idx).border = border
        row += 1

    row += 2
    ws.merge_cells(f"A{row}:P{row}")
    ws[f"A{row}"] = ("DOSÍMETRO DE CONTROL: incluido en cada paquete entregado para monitorear la exposición a la radiación "
                     "recibida durante el tránsito y almacenamiento. Este dosímetro debe ser guardado por el cliente en un "
                     "área libre de radiación durante el período de uso.")
    ws[f"A{row}"].font = Font(size=10, bold=True); ws[f"A{row}"].alignment = Alignment(horizontal="left", wrap_text=True)
    row += 2

    ws.merge_cells(f"A{row}:P{row}")
    ws[f"A{row}"] = ("POR DEBAJO DEL MÍNIMO DETECTADO: es la dosis por debajo de la cantidad mínima reportada para el período "
                     "de uso y son registradas como \"PM\".")
    ws[f"A{row}"].font = Font(size=10, bold=True); ws[f"A{row}"].alignment = Alignment(horizontal="left", wrap_text=True)

    bio = BytesIO(); wb.save(bio); bio.seek(0)
    return bio.getvalue()

# Preparar logo/encabezado
header_lines = [header_line1, header_line2, header_line3, header_line4]
logo_bytes = logo_file.read() if logo_file is not None else None

xlsx_fmt = build_formatted_excel(out.copy(), header_lines, logo_bytes)
st.download_button(
    "⬇️ Descargar Excel (formato plantilla)",
    data=xlsx_fmt,
    file_name=f"reporte_dosimetria_plantilla_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)





